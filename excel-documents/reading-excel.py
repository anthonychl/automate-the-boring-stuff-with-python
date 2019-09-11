'''
an Excel document is also called a Workbook
a Workbook can contain one or more sheets/spreadsheets/worksheets (they can be called either way)
a sheet contains columns and rows and their intersection is called a cell

to work with excel docs we use openpyxl
which is a third party module so we must pip install it
pip install openpyxl
'''
import openpyxl

workbook = openpyxl.load_workbook('example.xlsx')
print( type(workbook)) # <class 'openpyxl.workbook.workbook.Workbook'>

'''
# this way of creating a sheet object is apparently deprecated
sheet = workbook.get_sheet_by_name('Sheet1')
print( type(sheet)) # <class 'openpyxl.worksheet.worksheet.Worksheet'>

# we should do it as its shown below
'''
sheet = workbook['Sheet1']
print( type(sheet)) # # <class 'openpyxl.worksheet.worksheet.Worksheet'>

'''
# what if we dont know the name(s) of the sheet(s) in the workbook?
# this is also deprecated
sh_list = workbook.get_sheet_names()
print(sh_list) # ['Sheet1', 'Sheet2', 'Sheet3']

#do it as shown below
'''
sh_list = workbook.sheetnames
print(sh_list) # ['Sheet1', 'Sheet2', 'Sheet3']

cell = sheet['A1'] # cell object = column 'A' row '1'
print(cell.value) # 2015-04-05 13:34:02  that cell contains datetime values
# another way without creating a cell object
print(sheet['B1'].value) # Apples
print(sheet['C1'].value) # 73

# another way without using letters but calling columns and rows by number
print(sheet.cell(row=1, column=2).value) # Apples
print(sheet.cell(row=1, column=3).value) # 73

# this sheet.cell(row=1, column=3) is the same as sheet['C1'] they return the same cell object
# this syntax while more complicated, could be useful if we need to iterate over the cells
print('---------------------------------')
for i in range(1, 8):
    for j in range(1, 4):
        print(sheet.cell(row=i, column=j).value)
    print('----------------------------')