import openpyxl
wb = openpyxl.Workbook() # create a new document
sh_list = wb.sheetnames # there should be only one and by default is 'Sheet'

sheet = wb[sh_list[0]] # or sheet = wb['Sheet']

# assign value to the cells
sheet['A1'] = 'Anthony'
sheet['B1'] = 'Chaple'

# now this document only exists for as long as our program is running
# if we want to save it to the HDD
wb.save('myname.xlsx') # this will save the doc to the current working dir

sheet2 = wb.create_sheet() # 'Sheet2'
print(wb.sheetnames)
# say we want to name a sheet differently
sheet2.title = 'NewSheet'
print(wb.sheetnames)

wb.save('myname2.xlsx') 
'''
if we are tinkering around with our code is better 
if we save the changes in a new doc
in case we mess up we'll still have the original excel
'''

sheet3 = wb.create_sheet(index=0, title='myOtherSheet') # add this sheet at the begining, index 0, we can insert them where we want
sheet4 = wb.create_sheet(index=2,title='thirdSheet') # we CAN NOT do this wb.create_sheet(2,'thirdSheet')
wb.save('myname3.xlsx')